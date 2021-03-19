from openpyxl import Workbook
from openpyxl import load_workbook
import os
from selenium import webdriver
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.webdriver import WebDriver


# class ChromeWebdriver:

def headless():
    # To Run the Program headless without opening any physical window

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.binary_location = '.\\Canary\\Chrome SxS\\Application\\chrome.exe'
    driver = webdriver.Chrome(options=chrome_options, executable_path='chromedriver83.exe')
    return driver


def normal():
    chrome_options = Options()
    chrome_options.binary_location = '.\\Canary\\Chrome SxS\\Application\\chrome.exe'
    driver = webdriver.Chrome(options=chrome_options, executable_path='chromedriver83.exe')
    # driver = webdriver.Chrome('chromedriver87.exe')
    return driver


def open_sheets():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Mambu Id"
    sheet["B1"] = "Amount"
    sheet["C1"] = "Link"
    sheet["D1"] = "Status"
    sheet["E1"] = "=Counta(A:A)"
    workbook.save(filename="LinkData.xlsx")
    os.startfile("LinkData.xlsx")


def logic(driver):
    driver.implicitly_wait(100)
    driver.get("https://prodsupport.zestmoney.in")
    driver.find_element_by_id("email-id").send_keys("lavanya@zestmoney.in")
    driver.find_element_by_id("pwd").send_keys("GRBhd38h")
    driver.find_element_by_class_name("jss175").click()
    time.sleep(5)
    driver.get("https://prodsupport.zestmoney.in/dashboard/PaymentLink")

    wb = load_workbook(r'.\LinkData.xlsx')
    sheet = wb.active
    # To check the count of the Mambu id to control the Loop
    z = sheet.max_row
    # z = sheet.cell(1, 5).value
    # z = int(counta)
    # Initialize
    a = 2

    # for loop to check and tell the program to stop

    while a <= z:
        # if loop to check the id are already generated?

        if sheet.cell(a, 4).value != 'Done':
            # Getting Mambu id and putting in the website.
            MAMBU_ID = sheet.cell(a, 1).value
            driver.find_element_by_xpath(
                '/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[1]/div/input').clear()
            search_input1 = driver.find_element_by_xpath(
                '/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[1]/div/input')
            search_input1.send_keys(MAMBU_ID)

            # Getting Amount and putting in the website.
            AMOUNT = sheet.cell(a, 2).value
            driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[2]/div/input").clear()
            search_input2 = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[2]/div/input")
            search_input2.send_keys(AMOUNT)

            # click to generate Link
            driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[3]/div/div/select/option[3]").click()
            driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div[4]/button/span[1]").click()

            # wait for the Link to generate just in case poor connection
            time.sleep(5)

            # check the Link is generated or its a error
            suxes = driver.find_element_by_xpath("/html/body/div[4]/div[2]/div/div[1]/h6").text
            err = driver.find_element_by_xpath("/html/body/div[4]/div[2]/div/div[1]/h6").text

            if suxes == 'SUCCESS':
                link = driver.find_element_by_xpath("/html/body/div[4]/div[2]/div/div[2]/div[2]").text

            elif err == 'Error':
                link = driver.find_element_by_xpath("/html/body/div[4]/div[2]/div/div[2]/div").text

            # ---------------------------------#----------------------------------------
            try:
                # update the sheet with the link
                sheet.cell(a, 3).value = link
                # link_obj = link

                # update the sheet with completion status

                sheet.cell(a, 4).value = 'Done'

                # click cancel button
                driver.find_element_by_xpath("/html/body/div[4]/div[2]/div/div[3]/button").click()

                # iterate
                a += 1
            finally:
                wb.save(filename="LinkData.xlsx")


        else:
            a += 1
            wb.save(filename="LinkData.xlsx")

    wb.save(filename="LinkData.xlsx")
    driver.quit()
